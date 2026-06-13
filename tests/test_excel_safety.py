from io import BytesIO

import pandas as pd
from openpyxl import load_workbook

from main import _excel_safe_cell, frames_to_excel_bytes


def test_excel_safe_cell_escapes_formula_triggers():
    assert _excel_safe_cell("=1+2") == "'=1+2"
    assert _excel_safe_cell("@SUM(A1)") == "'@SUM(A1)"
    assert _excel_safe_cell("+cmd") == "'+cmd"
    assert _excel_safe_cell("-cmd") == "'-cmd"


def test_excel_safe_cell_keeps_legit_numbers_and_non_strings():
    assert _excel_safe_cell("-3.21%") == "-3.21%"  # 負百分比（- 後接數字）不動
    assert _excel_safe_cell("+886912") == "+886912"
    assert _excel_safe_cell("normal") == "normal"
    assert _excel_safe_cell("") == ""
    assert _excel_safe_cell(123) == 123
    assert _excel_safe_cell(None) is None


def test_frames_to_excel_bytes_writes_escaped_workbook():
    frames = {
        "pass": pd.DataFrame(
            {
                "股票名稱": ["=HYPERLINK(\"http://evil\")", "正常股"],
                "近3個月漲幅": ["-3.21%", "+5%"],
            }
        )
    }
    workbook = load_workbook(BytesIO(frames_to_excel_bytes(frames)))
    sheet = workbook["pass"]
    # 第 1 列為表頭，資料自第 2 列起
    assert sheet["A2"].value == "'=HYPERLINK(\"http://evil\")"  # 公式被轉義
    assert sheet["A3"].value == "正常股"
    assert sheet["B2"].value == "-3.21%"  # 合法負百分比不被改動


def test_frames_to_excel_bytes_handles_empty_frame():
    workbook = load_workbook(BytesIO(frames_to_excel_bytes({"pass": pd.DataFrame()})))
    assert "pass" in workbook.sheetnames
